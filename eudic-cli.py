#!/usr/bin/env python3

import sys
import unicodedata

import click
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup


def half2full(s):
    ret = []
    for char in s:
        if unicodedata.east_asian_width(char) in ("F", "W"):
            ret.append(char)
        elif unicodedata.category(char) == "Mn":
            ret.append(unicodedata.normalize("NFKC", char))
        else:
            ret.append(unicodedata.normalize("NFKC", char))
    return "".join(ret)


@click.group()
def cli():
    pass


@cli.command("excel")
@click.argument("input", type=click.File("r"), default=sys.stdin)
@click.argument("output", type=click.File("a+"), default=sys.stdout)
@click.option("-c", "--chapter", type=int, default=1)
def to_excel(input, output, chapter):
    html = input.read()
    soup = BeautifulSoup(html, "html.parser")
    data = {"words": []}
    for tr in soup.find_all("tr")[1:]:
        td_list = tr.find_all("td")
        word = td_list[1].text.strip()
        data["words"].append(word)

    df = pd.DataFrame(data)
    df["phrases"] = ""
    df["examples"] = ""
    click.secho(f"Found {df.count()} words", fg="green")
    if output.name == "<stdout>":
        print(df)
    else:
        with pd.ExcelWriter(output.name, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=f"Chapter {chapter:02d}")


@cli.command("txt")
@click.argument("input", type=click.Path(exists=True, file_okay=True))
@click.argument("output", type=click.File("w"), default=sys.stdout)
@click.option("-c", "--chapter", type=int, default=0)
def to_txt(input, output, chapter):
    assert chapter >= 0
    workbook = openpyxl.load_workbook(input)
    dfs = []
    if chapter == 0:
        sheets = workbook.sheetnames
    else:
        sheets = [workbook.sheetnames[chapter - 1]]
    for sheet in sheets:
        df = pd.read_excel(input, sheet_name=sheet, engine="openpyxl")
        dfs.append(df)

    df = pd.concat(dfs).fillna("")
    df["phrases"] = df["phrases"].apply(half2full)
    df["examples"] = df["examples"].apply(half2full)
    if output.name == "<stdout>":
        print(df)
    else:
        records = df.to_dict(orient="records")
        content = []
        for x in records:
            word = x["words"]
            phrases = "<br>".join(x["phrases"].split("\n"))
            examples = "<br>".join(x["examples"].split("\n"))
            descriptions = [x for x in [phrases, examples] if x]
            descriptions = "<br>".join(descriptions)
            content.append(f"{word}@{descriptions}\n")
        output.writelines(content)


if __name__ == "__main__":
    cli()
